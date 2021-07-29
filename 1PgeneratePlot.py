'''
One-Round Group Testing Algorithm

n = number of people to be tested
p = prevalence (estimated fraction of infected people in our test group)
p = k / n where k is the number of infected people
s = group size
partition = number of partitions to perform

S = the set of people to be tested
Form partition partitions of S into n/s groups of s people (all together there are rn/s groups)
Test every group
for each person x
     if x is in a group that tested negative then
    report that x is negative
    remove x from all groups
for each person x
     if x is the only person left in a group that tested positive
           report that x is positive
'''
import math
import random
from collections import defaultdict
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from matplotlib.colors import Normalize
from statistics import mean,variance
import openpyxl

wb = openpyxl.Workbook()

WORD_SIZE_BYTES = 8
from os import path
import sys



def repToInt(arr, base):
    llen = len(arr)
    power = 1
    num = 0

    # int equivalent is arr[len-1]*1 +
    # arr[len-2]*base + arr[len-3]*(base^2) + ...
    for i in range(llen - 1, -1, -1):
        if int(arr[i]) >= base:
            print('Invalid Number passed to repToDec:' + arr[i])
            exit(1)
        num += int(arr[i]) * power
        power = power * base
    return num


def intToRep(inputNum, base):
    # Convert input number to given base
    # by repeatedly dividing it by base
    # and taking remainder
    if inputNum == 0:
        return [0]
    result = list()
    while (inputNum > 0):
        result.append(inputNum % base)
        inputNum = inputNum // base
    # Reverse the result
    result = result[::-1]

    return result

class TestGroup:
    def __init__(self):
        self.set = set()
        self.testsPositive = None


if len(sys.argv) < 2:
    print("Input should be: generateIsoClasses.py testdata.bin")
    exit(1)
if not path.exists(sys.argv[1]):
    print("File not found: " + sys.argv[1])
    exit(1)


idPositive = set()
idNegative = set()
groups = list()
newPositives = set()


testfile = open(sys.argv[1], "rb")
params=list()
params.append(int.from_bytes(testfile.read(2), "big"))
params.append(int.from_bytes(testfile.read(2), "big"))
params.append(int.from_bytes(testfile.read(2), "big"))
params.append(int.from_bytes(testfile.read(2), "big"))
n = int(params[0])


params.clear()

nextWord_test = testfile.read(WORD_SIZE_BYTES)

row = list()
testdata = []
while nextWord_test:
    nextInt = int.from_bytes(nextWord_test, "big")
    nextBaseRep = intToRep(nextInt, 2)
    nextWord_test = testfile.read(WORD_SIZE_BYTES)
    if nextWord_test == b'' and len(row)+len(nextBaseRep) < n:
        nextBaseRep = [0] * (n - len(nextBaseRep) - len(row)) + nextBaseRep
    elif len(nextBaseRep) < 64 and nextWord_test is not None:
        nextBaseRep = [0] * (64 - len(nextBaseRep)) + nextBaseRep
    row.extend(nextBaseRep)
    while len(row) > n:
        testdata.append(row[:n])
        row = row[n:len(row)]
    if len(row) == n:
        testdata.append(row[:])
        row.clear()

print("Test Data Loaded.")
testfile.close()
row.clear()
grouptests = []

k = testdata[0].count(1)
runs = 0
if len(sys.argv)>2:
    runs = int(sys.argv[2])
else:
    runs = len(testdata)

maxTests = n//2
output_filename = str(n)+"n_"+str(k)+"k_"+str(runs)+"runsPerGroupSize.xlsx"

ddPercentConfirmedPositive = defaultdict(list)
ddPercentConfirmedNegative = defaultdict(list)

ws_pos = wb['Sheet']
ws_pos.title = 'Plot Positive'
ws_neg = wb.create_sheet('Plot Negative')
ws_datapoints_pos = wb.create_sheet('Datapoints Positive')
ws_datapoints_neg = wb.create_sheet('Datapoints Negative')
ws_var = wb.create_sheet('Variance')
numDatapoints = 0

smin = n//(2*k) if n//(2*k) > 1 else 2
smax = (2*n)//k if (2*n)//k < n//2 else n//2

fig_pos = plt.figure()
fig_neg = plt.figure()

ax_pos = fig_pos.add_subplot(111)
ax_pos.set_ylabel('avg % confirmed positive')
ax_pos.set_xlabel('# of tests')
ax_neg = fig_neg.add_subplot(111)
ax_neg.set_ylabel('avg % confirmed negative')
ax_neg.set_xlabel('# of tests')

error_threshold = .0025

ws_var.cell(row=1, column=1, value="s")
ws_var.cell(row=1, column=2, value="max sample variance")
ws_var.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
ws_var.cell(row=1, column=4, value="m.s. var / sample size")
ws_var.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
ws_var.cell(row=1, column=6, value="std. dev")
ws_var.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
ws_var.cell(row=1, column=8, value="error threshold")
ws_var.cell(row=2, column=8, value=error_threshold)
ws_var.cell(row=1, column=9, value="num errors in 1 std. dev")
ws_var.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
ws_var.cell(row=1, column=11, value="P(an error)")
ws_var.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12)


for col in range(2,15):
    if col < 8:
        if col % 2 == 0:
            ws_var.cell(row=2, column=col, value="Positive")
        else:
            ws_var.cell(row=2, column=col, value="Negative")
    elif col > 8:
        if col % 2 == 1:
            ws_var.cell(row=2, column=col, value="Positive")
        else:
            ws_var.cell(row=2, column=col, value="Negative")

ws_datapoints_pos.cell(row=1, column=1, value="s")
ws_datapoints_neg.cell(row=1, column=1, value="s")
ws_datapoints_pos.cell(row=1, column=2, value="# tests")
ws_datapoints_neg.cell(row=1, column=2, value="# tests")
ws_datapoints_pos.cell(row=1, column=3, value="avg % id pos")
ws_datapoints_neg.cell(row=1, column=3, value="avg % id neg")
ws_datapoints_pos.cell(row=1, column=4, value="sample variance")
ws_datapoints_neg.cell(row=1, column=4, value="sample variance")

#list the s values to help normalize the color map
sList = list()
for size in range(smin,smax+1):
    if n%size==0:
        sList.append(size)
        ws_var.cell(row=2+len(sList), column=1, value=size)


norm_c = Normalize(vmin=0, vmax=len(sList) - 1)
cmap = cm.ScalarMappable(Normalize(0,1), 'brg').get_cmap()




for sIndex in range(len(sList)):
    s = sList[sIndex]
    ws_s = wb.create_sheet('s= '+str(sList[sIndex]))
    ws_s['A1'] = "x = run; y = partition"
    maxPartition = 0

    groupfilename = "grouptests_"+str(n)+"n_"+str(k)+"k_"+str(s)+"s_10000runs.bin"

    groupfile = open(groupfilename, "rb")

    params.append(int.from_bytes(groupfile.read(2), "big"))
    params.append(int.from_bytes(groupfile.read(2), "big"))
    params.append(int.from_bytes(groupfile.read(2), "big"))
    params.append(int.from_bytes(groupfile.read(2), "big"))
    g = int(params[1])
    maxValuesPerWord_groups = int(params[3])
    params.clear()
    row.clear()
    nextWord_group = groupfile.read(WORD_SIZE_BYTES)
    while nextWord_group:
        nextInt = int.from_bytes(nextWord_group, "big")
        nextBaseRep = intToRep(nextInt, g)
        nextWord_group = groupfile.read(WORD_SIZE_BYTES)
        if nextWord_group == b'' and len(row)+len(nextBaseRep) < n:
            nextBaseRep = [0] * (n - len(nextBaseRep) - len(row)) + nextBaseRep
        elif len(nextBaseRep) < maxValuesPerWord_groups and nextWord_group != b'':
            nextBaseRep = [0] * (maxValuesPerWord_groups - len(nextBaseRep)) + nextBaseRep
        row.extend(nextBaseRep)

        while len(row) > n:
            grouptests.append(row[:n])
            row = row[n:len(row)]
        if len(row) == n:
            grouptests.append(row[:])
            row.clear()
    groupfile.close()
    print(len(grouptests))
    print("Group Tests Loaded for s="+str(s))

    newGroups = list()
    groupTestsIndex = random.randint(0, len(grouptests)-1)
    ax_pos.scatter(x=0, y=0, color=cmap(norm_c(sIndex)),
                   label='s = ' + str(sList[sIndex]))
    ax_neg.scatter(x=0, y=0, color=cmap(norm_c(sIndex)),
                   label='s = ' + str(sList[sIndex]))
    for run in range(1,runs+1):
        print("run #" + str(run))
        groups.clear()
        partition = 1
        numRemaining = n
        ws_s.cell(row=1+run, column=1, value=run)

        while numRemaining > 0 and (n//s * partition) <= maxTests:
            # Form r partitions of S into n//s groups of s people (all together there are rn//s groups)
            for newGroup in range(g):
                newGroups.append(TestGroup())
            assignment = 0
            for i in range(n):
                if i not in idPositive and i not in idNegative:
                    try:
                        assignment = grouptests[groupTestsIndex][i]
                    except IndexError:
                        print("groupTestsIndex ="+str(groupTestsIndex)+", i = "+str(i)+", len(grouptests)="+str(len(grouptests)))
                        exit(1)
                    try:
                        newGroups[assignment].set.add(i)
                    except IndexError:
                        print(
                            "groupTestsIndex =" + str(groupTestsIndex) + ", i = " + str(i) + ", len(grouptests)=" + str(
                                len(grouptests))+", assignment="+str(assignment)+", len(newGroups)="+str(len(newGroups)))
                        exit(1)

            newGroups = [group for group in newGroups if len(group.set) > 0]

            # Test every new group
            for i in range(len(newGroups)):
                if any(testdata[run-1][j] == 1 for j in newGroups[i].set):
                    newGroups[i].testsPositive = True
                else:
                    newGroups[i].testsPositive = False

            groups.extend(newGroups)

            lenG = len(groups)
            for i in range(len(newGroups)):
                # if x is in a group that tested negative then
                # report that x is negative
                # remove x from all groups

                if groups[lenG - i - 1].testsPositive == False:
                    idNegative = idNegative.union(groups[lenG - i - 1].set)
                    groups.pop(lenG - i - 1)
            for group in groups:
                group.set = group.set.difference(idNegative)



            i = 0

            while i < len(groups):
                # if x is the only person left in a group that tested positive
                # report that x is positive
                # remove groups containing the positive (we're done with them)
                if len(groups[len(groups) - i - 1].set) == 1 and groups[len(groups) - i - 1].testsPositive:
                    for x in groups[len(groups) - i - 1].set:
                        newPositives.add(x)  # x is the lone member of groups[g]
                i += 1


            for pos in newPositives:
                i = len(groups)
                while i > 0:
                    i -= 1
                    if pos in groups[i].set or len(groups[i].set) == 0:
                        groups.pop(i)

            idPositive = idPositive.union(newPositives)
            newPositives.clear()
            newGroups.clear()

            #save counts by run, by partition
            if partition > maxPartition:
                maxPartition = partition
                ws_s.cell(row=1, column=1 + partition, value=partition)

            ws_s.cell(row=1+run,column=1+partition,value=groupTestsIndex+1)
            numRemaining = n - len(idPositive) - len(idNegative)
            ddPercentConfirmedPositive[partition].append(len(idPositive)/k)
            ddPercentConfirmedNegative[partition].append(len(idNegative)/(n-k))

            groupTestsIndex = random.randint(0, len(grouptests)-1)
            print(partition)
            partition += 1

        idPositive.clear()
        idNegative.clear()

    # do this after all runs of a valid s
    variances_pos = list()
    variances_neg = list()
    for r in ddPercentConfirmedNegative.keys():
        for fill in range(len(ddPercentConfirmedPositive[r]), runs):
            ddPercentConfirmedPositive[r].append(1)
        for fill in range(len(ddPercentConfirmedNegative[r]), runs):
            ddPercentConfirmedNegative[r].append(1)
        mean_pos = mean(ddPercentConfirmedPositive[r])
        mean_neg = mean(ddPercentConfirmedNegative[r])
        ax_pos.scatter(x=n//s * r, y=mean_pos, color=cmap(norm_c(sIndex)))
        ax_neg.scatter(x=n//s * r, y=mean_neg, color=cmap(norm_c(sIndex)))
        numDatapoints += 1
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=1, value=s)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=1, value=s)
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=2, value=n//s * r)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=2, value=n//s * r)
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=3, value=mean_pos)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=3, value=mean_neg)
        nextVar = variance(ddPercentConfirmedPositive[r])
        variances_pos.append(nextVar)
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=4, value=nextVar)
        nextVar = variance(ddPercentConfirmedNegative[r])
        variances_neg.append(nextVar)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=4, value=nextVar)

    for r in range(max(ddPercentConfirmedNegative, key=int), int(maxTests / (n//s))):
        numDatapoints += 1
        ax_pos.scatter(x=n // s * r, y=1, color=cmap(norm_c(sIndex)))
        ax_neg.scatter(x=n // s * r, y=1, color=cmap(norm_c(sIndex)))
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=1, value=s)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=1, value=s)
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=2, value=n//s * r)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=2, value=n//s * r)
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=3, value=1)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=3, value=1)
        nextVar = variance(ddPercentConfirmedPositive[r])
        variances_pos.append(nextVar)
        ws_datapoints_pos.cell(row=1 + numDatapoints, column=4, value=nextVar)
        nextVar = variance(ddPercentConfirmedNegative[r])
        variances_neg.append(nextVar)
        ws_datapoints_neg.cell(row=1 + numDatapoints, column=4, value=nextVar)

    maxVar = max(variances_pos)
    ws_var.cell(row=3+sIndex, column=2, value=maxVar)
    ws_var.cell(row=3 + sIndex, column=4, value=maxVar/runs)
    stddev = math.sqrt(maxVar / runs)
    ws_var.cell(row=3 + sIndex, column=6, value=stddev)
    ws_var.cell(row=3 + sIndex, column=9, value='=$H$2/F'+str(sIndex+3))
    ws_var.cell(row=3 + sIndex, column=11, value='=2*(1-_xlfn.NORM.DIST(I'+str(sIndex+3)+',0,1,TRUE))')

    maxVar = max(variances_neg)
    ws_var.cell(row=3 + sIndex, column=3, value=maxVar)
    ws_var.cell(row=3 + sIndex, column=5, value=maxVar / runs)
    stddev = math.sqrt(maxVar / runs)
    ws_var.cell(row=3 + sIndex, column=7, value=stddev)
    ws_var.cell(row=3 + sIndex, column=10, value='=$H$2/G'+str(sIndex+3))
    ws_var.cell(row=3 + sIndex, column=12, value='=2*(1-_xlfn.NORM.DIST(J'+str(sIndex+3)+',0,1,TRUE))')

    variances_pos.clear()
    variances_neg.clear()
    ddPercentConfirmedPositive.clear()
    ddPercentConfirmedPositive.clear()
    grouptests.clear()
ws_var.cell(row=1, column=13, value="P(at least one error in "+str(numDatapoints)+" datapoints)")
ws_var.merge_cells(start_row=1, start_column=13, end_row=1, end_column=14)
for sIndex in range(len(sList)):
    ws_var.cell(row=3 + sIndex, column=13, value="="+str(numDatapoints)+"*K"+str(sIndex+3))
    ws_var.cell(row=3 + sIndex, column=14, value="="+str(numDatapoints)+"*L"+str(sIndex+3))
testfile.close()



#do stuff with data

ax_pos.legend(loc='lower right')
ax_neg.legend(loc='lower right')
ax_pos.set_title("n = "+str(n)+", k = "+str(k))
fig_pos.savefig("pos.png", dpi=175)
img = openpyxl.drawing.image.Image('pos.png')
img.anchor='A1'
ws_pos.add_image(img)
ax_neg.set_title("n = "+str(n)+", k = "+str(k))
fig_neg.savefig("neg.png", dpi=175)
img = openpyxl.drawing.image.Image('neg.png')
img.anchor='A1'
ws_neg.add_image(img)
print("All runs complete. Scatter plot complete.")
wb.save(output_filename)
print(output_filename+" saved.")
exit(0)

