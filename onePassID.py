'''
One-Round Group Testing Algorithm

n = number of people to be tested
p = prevalence (estimated fraction of infected people in our test group)
p = k / n where k is the number of infected people
s = group size
r = number of partitions to perform

S = the set of people to be tested
Form r partitions of S into n/s groups of s people (all together there are rn/s groups)
Test every group
for each person x
     if x is in a group that tested negative then
    report that x is negative
    remove x from all groups
for each person x
     if x is the only person left in a group that tested positive
           report that x is positive
'''
import random
from scipy.stats.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()


class TestSubject:
    def __init__(self, id, hasCOVID19):
        self.id = id
        self.hasCOVID19 = hasCOVID19


class TestGroup:
    def __init__(self, id, set):
        self.id = id
        self.set = set
        self.testsPositive = None

def purge(purgeMember, allGroups):
        #purges groups containing a certain member
        range = len(allGroups)
        i=range - 1
        while i >=0 :
            if purgeMember in allGroups[i].set:
                allGroups.remove(i)
                i-=1

def corellateMembershipWithInfectionStatus(allGroups):
    #returns corellation value for infection status versus how many groups each person is in, post-elimination
    members = dict()
    for group in allGroups:
        for person in group.set:
            members[person] = members.get(person,0) + 1
    infectionStatuses = [1 if member.hasCOVID19 else 0 for member in members.keys()]
    membership = list(members.values())
    if len(members) > 1:
        corr = pearsonr(infectionStatuses, membership)[0]
        return corr
    else:
        return None


n = 720
k = 8
s = 90

S = [TestSubject(x, True) if x < k else TestSubject(x, False) for x in range(n)]

r = 0
nextPartition = list()
idPositive = set()
idNegative = set()
groups = list()
valuessheet = wb['Sheet']
valuessheet.title = 'Values'
valuessheet.cell(row=1, column=1, value="Partitions")
valuessheet.cell(row=1, column=2, value="ID Positives")
valuessheet.cell(row=1, column=3, value="ID Negatives")
valuessheet.cell(row=1, column=4, value="Negative Inconclusives")
valuessheet.cell(row=1, column=5, value="Positive Inconclusives")
valuessheet.cell(row=1, column=6, value="Correlation(Infection Rate/Membership in Unconfirmed Groups)")

newPositives = set()
newPositives.clear()

while len(idPositive) + len(idNegative) < n:
    # Form r partitions of S into n//s groups of s people (all together there are rn//s groups)

    r += 1
    sheetName = str(r)+" partition(s)"
    ws = wb.create_sheet(sheetName)

    random.shuffle(S)
    nextPartition = S.copy()
    #add groups from new partition, cull those confirmed
    for g in range(n//s):
        nextSet = set(
            nextPartition[(g*s): ((g+1)*s)]).difference(idNegative.union(idPositive))
        if len(nextSet)>0:
            groups.append(TestGroup(str(r)+"-"+str(g), nextSet))


    for g in range(len(groups)):
        i=0
        for person in groups[g].set:
            i+=1
            cell = ws.cell(row=g+2, column=i, value=person.id)
            if person.hasCOVID19:
                cell.font = Font(bold=True)

    # Test every group
    for g in range(len(groups)):
        if any(person.hasCOVID19 for person in groups[g].set):
            groups[g].testsPositive = True
            cell = ws.cell(row=g+2, column=s + 1, value="(+) Positive")
        else:
            groups[g].testsPositive = False
            cell = ws.cell(row=g+2, column=s + 1, value="(-) Negative")

    cell = ws.cell(row=1, column=s+3, value="ID/PURGE NEGATIVES")
    for g in range(len(groups)):
        # if x is in a group that tested negative then
        # report that x is negative
        # remove x from all groups
        if groups[g].testsPositive==False:
            idNegative = idNegative.union(groups[g].set)
            purgeSet = groups[g].set.copy()
            for group in groups:
                group.set = group.set.difference(purgeSet)

    for g in range(len(groups)):
        col = s+2
        for person in groups[g].set:
            col+=1
            cell = ws.cell(row=g + 2, column=col, value=person.id)
            if person.hasCOVID19:
                cell.font = Font(bold=True)

    numGroups = len(groups)
    g=0

    while g<numGroups:
        # if x is the only person left in a group that tested positive
        # report that x is positive
        # remove groups containing the positive (we're done with them)
        if groups[g].testsPositive and len(groups[g].set) == 1:

            for x in groups[g].set:
                myPos = x
            #myPos is the unique element of groups[g].set
            newPositives.add(myPos)
        g+=1

    for pos in newPositives:
        i = len(groups)
        while i > 0:
            i-=1
            if pos in groups[i].set:
                groups.pop(i)

    idPositive=idPositive.union(newPositives)
    newPositives.clear()

    cell = ws.cell(row=1, column=2*s+4, value="ID SINGLETON POSITIVES\PURGE GROUPS CONTAINING THEM")
    for g in range(len(groups)):
        col = 2*s+3
        for person in groups[g].set:
            col += 1
            cell = ws.cell(row=g + 2, column=col, value=person.id)
            if person.hasCOVID19:
                cell.font = Font(bold=True)

    groups = [group for group in groups if len(group.set) > 0]

    remaining = n - (len(idPositive) + len(idNegative))
    print('For', r, 'partition(s) I found', len(idPositive), 'positives and', len(idNegative), 'negatives.', remaining, 'remain(s) unconfirmed.')
    valuessheet.cell(row=r+1, column=1, value=r)
    valuessheet.cell(row=r+1, column=2, value=len(idPositive))
    valuessheet.cell(row=r+1, column=3, value=len(idNegative))
    valuessheet.cell(row=r+1, column=4, value=n-k-len(idNegative))
    valuessheet.cell(row=r+1, column=5, value=k-len(idPositive))
    valuessheet.cell(row=r + 1, column=6, value=corellateMembershipWithInfectionStatus(groups))
wb.save("infographic.xlsx")
